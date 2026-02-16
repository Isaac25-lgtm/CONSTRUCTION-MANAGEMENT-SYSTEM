import csv
import io
from datetime import date, timedelta
from decimal import Decimal
from typing import Any
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from app.api.v1.dependencies import (
    OrgContext,
    ensure_project_permission,
    get_org_context,
    get_project_or_404,
)
from app.db.session import get_db
from app.models.boq import BOQHeader, BOQItem
from app.models.task import Task, TaskPriority, TaskStatus
from app.schemas.boq import (
    BOQHeaderResponse,
    BOQImportResponse,
    BOQItemResponse,
    BOQItemUpdate,
    BOQSummaryResponse,
)

router = APIRouter()


def _decimal_value(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value))
    except Exception:
        return default


def _int_value(value: Any, default: int = 0, min_value: int | None = None, max_value: int | None = None) -> int:
    if value is None or value == "":
        parsed = default
    else:
        try:
            parsed = int(value)
        except Exception:
            parsed = default
    if min_value is not None:
        parsed = max(min_value, parsed)
    if max_value is not None:
        parsed = min(max_value, parsed)
    return parsed


def _compute_budget(quantity: Decimal, rate: Decimal) -> Decimal:
    return (quantity * rate).quantize(Decimal("0.01"))


def _compute_variance(budget_cost: Decimal, actual_cost: Decimal) -> Decimal:
    return (budget_cost - actual_cost).quantize(Decimal("0.01"))


def _load_rows_from_upload(file: UploadFile) -> list[dict[str, Any]]:
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        raw = file.file.read()
        decoded = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        return [dict(row) for row in reader]

    if filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="XLSX import requires openpyxl runtime dependency",
            ) from exc
        raw = file.file.read()
        workbook = load_workbook(filename=io.BytesIO(raw), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        parsed: list[dict[str, Any]] = []
        for row in rows[1:]:
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[idx] if idx < len(row) else None
            parsed.append(item)
        return parsed

    raise HTTPException(status_code=400, detail="Only CSV and XLSX BOQ imports are supported")


def _leaf_items(items: list[BOQItem]) -> list[BOQItem]:
    parent_ids = {item.parent_item_id for item in items if item.parent_item_id}
    return [item for item in items if item.id not in parent_ids]


def _compute_weighted_completion(items: list[BOQItem]) -> float:
    leaves = _leaf_items(items)
    if not leaves:
        return 0.0

    total_weight = Decimal("0")
    weighted_sum = Decimal("0")
    for item in leaves:
        weight_fraction = Decimal(item.weight_out_of_10 or 0) / Decimal("10")
        completion_fraction = Decimal(item.percent_complete or 0) / Decimal("100")
        total_weight += weight_fraction
        weighted_sum += weight_fraction * completion_fraction

    if total_weight <= 0:
        return float(
            sum((item.percent_complete or 0) for item in leaves) / max(len(leaves), 1)
        )

    return float((weighted_sum / total_weight) * Decimal("100"))


def _serialize_item(item: BOQItem) -> BOQItemResponse:
    return BOQItemResponse(
        id=item.id,
        header_id=item.header_id,
        organization_id=item.organization_id,
        project_id=item.project_id,
        parent_item_id=item.parent_item_id,
        item_code=item.item_code,
        description=item.description,
        unit=item.unit,
        quantity=item.quantity,
        rate=item.rate,
        budget_cost=item.budget_cost,
        weight_out_of_10=item.weight_out_of_10,
        percent_complete=item.percent_complete,
        actual_cost=item.actual_cost,
        variance=_compute_variance(item.budget_cost, item.actual_cost),
        created_at=item.created_at,
        updated_at=item.updated_at,
        children=[],
    )


def _build_item_tree(items: list[BOQItem]) -> list[BOQItemResponse]:
    mapped = {item.id: _serialize_item(item) for item in items}
    roots: list[BOQItemResponse] = []
    for item in items:
        node = mapped[item.id]
        if item.parent_item_id and item.parent_item_id in mapped:
            mapped[item.parent_item_id].children.append(node)
        else:
            roots.append(node)
    return roots


def _get_or_create_header(
    db: Session,
    *,
    org_id: UUID,
    project_id: UUID,
    title: str,
    start_date: date | None = None,
    end_date: date | None = None,
    currency: str = "UGX",
) -> BOQHeader:
    header = (
        db.query(BOQHeader)
        .filter(BOQHeader.organization_id == org_id, BOQHeader.project_id == project_id)
        .order_by(BOQHeader.updated_at.desc())
        .first()
    )
    if header:
        header.title = title
        header.start_date = start_date
        header.end_date = end_date
        header.currency = currency
        db.query(BOQItem).filter(BOQItem.header_id == header.id).delete(synchronize_session=False)
        db.flush()
        return header

    header = BOQHeader(
        organization_id=org_id,
        project_id=project_id,
        title=title,
        start_date=start_date,
        end_date=end_date,
        currency=currency,
    )
    db.add(header)
    db.flush()
    return header


def _get_latest_header_or_404(db: Session, *, org_id: UUID, project_id: UUID) -> BOQHeader:
    header = (
        db.query(BOQHeader)
        .filter(BOQHeader.organization_id == org_id, BOQHeader.project_id == project_id)
        .order_by(BOQHeader.updated_at.desc())
        .first()
    )
    if not header:
        raise HTTPException(status_code=404, detail="No BOQ found for this project")
    return header


@router.post("/import", response_model=BOQImportResponse)
async def import_boq(
    project_id: UUID,
    file: UploadFile = File(...),
    title: str = Query("Project BOQ"),
    currency: str = Query("UGX"),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
    ctx: OrgContext = Depends(get_org_context),
    db: Session = Depends(get_db),
):
    project = get_project_or_404(db, ctx.organization.id, project_id)
    ensure_project_permission(
        db,
        project,
        ctx.user,
        "can_edit_tasks",
        "You do not have permission to import BOQ data for this project",
    )

    rows = _load_rows_from_upload(file)
    if not rows:
        raise HTTPException(status_code=400, detail="Import file has no data rows")

    header = _get_or_create_header(
        db,
        org_id=ctx.organization.id,
        project_id=project_id,
        title=title,
        start_date=start_date,
        end_date=end_date,
        currency=currency,
    )

    created_items: list[BOQItem] = []
    parent_code_map: dict[str, str] = {}
    for row in rows:
        item_code = str(row.get("item_code", "")).strip() or None
        description = str(row.get("description", "")).strip()
        if not description:
            continue

        quantity = _decimal_value(row.get("quantity"), Decimal("0"))
        rate = _decimal_value(row.get("rate"), Decimal("0"))
        budget_cost = _compute_budget(quantity, rate)
        weight_out_of_10 = _int_value(row.get("weight_out_of_10"), default=1, min_value=0, max_value=10)
        percent_complete = _int_value(row.get("percent_complete"), default=0, min_value=0, max_value=100)
        actual_cost = _decimal_value(row.get("actual_cost"), Decimal("0"))

        item = BOQItem(
            header_id=header.id,
            organization_id=ctx.organization.id,
            project_id=project_id,
            item_code=item_code,
            description=description,
            unit=str(row.get("unit", "")).strip() or None,
            quantity=quantity,
            rate=rate,
            budget_cost=budget_cost,
            weight_out_of_10=weight_out_of_10,
            percent_complete=percent_complete,
            actual_cost=actual_cost,
            parent_item_id=None,
        )
        db.add(item)
        db.flush()
        created_items.append(item)

        parent_code = str(row.get("parent_item_code", "")).strip()
        if item_code and parent_code:
            parent_code_map[item_code] = parent_code

    by_code = {item.item_code: item for item in created_items if item.item_code}
    for item in created_items:
        if not item.item_code:
            continue
        parent_code = parent_code_map.get(item.item_code)
        if not parent_code:
            continue
        parent = by_code.get(parent_code)
        if parent and parent.id != item.id:
            item.parent_item_id = parent.id

    db.commit()
    return BOQImportResponse(header_id=header.id, imported_items=len(created_items))


@router.get("", response_model=BOQHeaderResponse)
async def get_boq(
    project_id: UUID,
    ctx: OrgContext = Depends(get_org_context),
    db: Session = Depends(get_db),
):
    project = get_project_or_404(db, ctx.organization.id, project_id)
    ensure_project_permission(
        db,
        project,
        ctx.user,
        "can_view_project",
        "You do not have permission to view BOQ data in this project",
    )

    header = _get_latest_header_or_404(db, org_id=ctx.organization.id, project_id=project_id)
    items = db.query(BOQItem).filter(BOQItem.header_id == header.id).all()
    return BOQHeaderResponse(
        id=header.id,
        organization_id=header.organization_id,
        project_id=header.project_id,
        title=header.title,
        start_date=header.start_date,
        end_date=header.end_date,
        currency=header.currency,
        created_at=header.created_at,
        updated_at=header.updated_at,
        items=_build_item_tree(items),
    )


@router.patch("/items/{item_id}", response_model=BOQItemResponse)
async def patch_boq_item(
    project_id: UUID,
    item_id: UUID,
    item_data: BOQItemUpdate,
    ctx: OrgContext = Depends(get_org_context),
    db: Session = Depends(get_db),
):
    project = get_project_or_404(db, ctx.organization.id, project_id)
    ensure_project_permission(
        db,
        project,
        ctx.user,
        "can_edit_tasks",
        "You do not have permission to update BOQ items in this project",
    )

    item = (
        db.query(BOQItem)
        .filter(
            BOQItem.id == item_id,
            BOQItem.organization_id == ctx.organization.id,
            BOQItem.project_id == project_id,
        )
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="BOQ item not found")

    update_data = item_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(item, field, value)

    item.weight_out_of_10 = _int_value(item.weight_out_of_10, default=1, min_value=0, max_value=10)
    item.percent_complete = _int_value(item.percent_complete, default=0, min_value=0, max_value=100)
    item.quantity = _decimal_value(item.quantity, Decimal("0"))
    item.rate = _decimal_value(item.rate, Decimal("0"))
    item.actual_cost = _decimal_value(item.actual_cost, Decimal("0"))
    item.budget_cost = _compute_budget(item.quantity, item.rate)

    db.commit()
    db.refresh(item)
    return _serialize_item(item)


@router.get("/summary", response_model=BOQSummaryResponse)
async def get_boq_summary(
    project_id: UUID,
    ctx: OrgContext = Depends(get_org_context),
    db: Session = Depends(get_db),
):
    project = get_project_or_404(db, ctx.organization.id, project_id)
    ensure_project_permission(
        db,
        project,
        ctx.user,
        "can_view_project",
        "You do not have permission to view BOQ summary in this project",
    )

    header = _get_latest_header_or_404(db, org_id=ctx.organization.id, project_id=project_id)
    items = db.query(BOQItem).filter(BOQItem.header_id == header.id).all()
    if not items:
        return BOQSummaryResponse(
            header_id=header.id,
            project_id=project_id,
            total_items=0,
            total_budget_cost=Decimal("0.00"),
            total_actual_cost=Decimal("0.00"),
            total_variance=Decimal("0.00"),
            project_weighted_completion_percent=0.0,
        )

    total_budget = sum((item.budget_cost for item in items), Decimal("0"))
    total_actual = sum((item.actual_cost for item in items), Decimal("0"))
    total_variance = _compute_variance(total_budget, total_actual)
    weighted_completion = _compute_weighted_completion(items)

    return BOQSummaryResponse(
        header_id=header.id,
        project_id=project_id,
        total_items=len(items),
        total_budget_cost=total_budget.quantize(Decimal("0.01")),
        total_actual_cost=total_actual.quantize(Decimal("0.01")),
        total_variance=total_variance,
        project_weighted_completion_percent=round(weighted_completion, 2),
    )


@router.post("/sync-tasks")
async def sync_tasks_from_boq(
    project_id: UUID,
    overwrite_existing: bool = Query(False),
    ctx: OrgContext = Depends(get_org_context),
    db: Session = Depends(get_db),
):
    project = get_project_or_404(db, ctx.organization.id, project_id)
    ensure_project_permission(
        db,
        project,
        ctx.user,
        "can_edit_tasks",
        "You do not have permission to sync tasks from BOQ in this project",
    )

    header = _get_latest_header_or_404(db, org_id=ctx.organization.id, project_id=project_id)
    items = db.query(BOQItem).filter(BOQItem.header_id == header.id).all()
    leaves = _leaf_items(items)
    if not leaves:
        raise HTTPException(status_code=400, detail="No BOQ leaf items to sync")

    existing_tasks = (
        db.query(Task)
        .filter(
            Task.organization_id == ctx.organization.id,
            Task.project_id == project_id,
            Task.is_deleted == False,
        )
        .all()
    )
    existing_by_name = {task.name: task for task in existing_tasks}

    synced = 0
    created = 0
    today = date.today()
    for item in leaves:
        task_name = item.description[:255]
        task = existing_by_name.get(task_name)
        due_date = header.end_date or (today + timedelta(days=30))
        start_date = header.start_date or today
        progress = _int_value(item.percent_complete, default=0, min_value=0, max_value=100)
        status = TaskStatus.COMPLETED if progress >= 100 else (
            TaskStatus.IN_PROGRESS if progress > 0 else TaskStatus.PENDING
        )

        if task and not overwrite_existing:
            task.progress = progress
            task.status = status
            task.start_date = task.start_date or start_date
            task.due_date = task.due_date or due_date
            synced += 1
            continue

        if task and overwrite_existing:
            task.description = f"Synced from BOQ item {item.item_code or ''}".strip()
            task.priority = TaskPriority.MEDIUM
            task.status = status
            task.progress = progress
            task.start_date = start_date
            task.due_date = due_date
            synced += 1
            continue

        new_task = Task(
            organization_id=ctx.organization.id,
            project_id=project_id,
            name=task_name,
            description=f"Synced from BOQ item {item.item_code or ''}".strip(),
            status=status,
            priority=TaskPriority.MEDIUM,
            reporter_id=ctx.user.id,
            assignee_id=None,
            start_date=start_date,
            due_date=due_date,
            progress=progress,
            dependencies=[],
        )
        db.add(new_task)
        created += 1

    db.commit()
    return {"synced": synced, "created": created, "leaf_items": len(leaves)}
